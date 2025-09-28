import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import sqlite3
import datetime
import calendar
import random
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from PIL import Image, ImageTk
from tkcalendar import DateEntry # DatePicker import

# --- 1. SQLite DB 및 스케줄 생성 로직 (이전과 동일) ---
DB_NAME = 'employees.db'
# ... (init_db, add_employee_db 등 모든 DB 및 스케줄 생성 함수들은 이전과 동일)
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_type TEXT NOT NULL,
            name TEXT NOT NULL,
            hire_date TEXT,
            contract_period TEXT,
            memo TEXT
        )
    ''')
    conn.commit()
    conn.close()

def add_employee_db(contract_type, name, hire_date, period, memo):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO employees (contract_type, name, hire_date, contract_period, memo) VALUES (?, ?, ?, ?, ?)",
                       (contract_type, name, hire_date, period, memo))
        conn.commit()
        return True
    except sqlite3.Error as e:
        messagebox.showerror("DB 오류", f"직원 추가 실패: {e}")
        return False
    finally:
        conn.close()

def get_all_employees_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM employees ORDER BY name")
    employees = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return employees

def update_employee_db(employee_id, contract_type, name, hire_date, period, memo):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE employees SET contract_type=?, name=?, hire_date=?, contract_period=?, memo=? WHERE id=?",
                       (contract_type, name, hire_date, period, memo, employee_id))
        conn.commit()
        return True
    except sqlite3.Error as e:
        messagebox.showerror("DB 오류", f"직원 수정 실패: {e}")
        return False
    finally:
        conn.close()

def delete_employee_db(employee_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM employees WHERE id=?", (employee_id,))
        conn.commit()
        return True
    except sqlite3.Error as e:
        messagebox.showerror("DB 오류", f"직원 삭제 실패: {e}")
        return False
    finally:
        conn.close()


MIN_WEEKDAY_STAFF = 5
MAX_WEEKDAY_STAFF = 6
WEEKEND_STAFF_COUNT = 6
START_ROW = 5

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
WORK_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
TITLE_FONT_EXCEL = Font(size=18, bold=True)
BOLD_FONT_EXCEL = Font(bold=True)
BLUE_FONT_EXCEL = Font(color="0000FF", bold=True)
RED_FONT_EXCEL = Font(color="FF0000", bold=True)
THIN_SIDE = Side(style="thin")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

def create_schedule_data(employees_data, year, month):
    schedule_data = defaultdict(set)
    work_days_required = {p['name']: int(p['contract_type'].replace('주 ', '').replace('회', '')) for p in employees_data}
    all_employee_names = [p['name'] for p in employees_data]
    weekend_shifts_assigned = {name: 0 for name in all_employee_names}

    _, num_days = calendar.monthrange(year, month)
    weeks = defaultdict(list)
    for day in range(1, num_days + 1):
        date = datetime.date(year, month, day)
        weeks[date.isocalendar()[1]].append(date)

    for week_num, dates_in_week in sorted(weeks.items()):
        work_days_scheduled_this_week = {name: 0 for name in all_employee_names}
        weekends_in_week = [d for d in dates_in_week if d.weekday() >= 5]
        weekdays_in_week = [d for d in dates_in_week if d.weekday() < 5]
        
        # 주말 근무 배정 (기존 로직 유지, WEEKEND_STAFF_COUNT가 최대 인원 역할을 함)
        for weekend_date in weekends_in_week:
            eligible = [name for name in all_employee_names if work_days_scheduled_this_week[name] < work_days_required[name]]
            if not eligible: continue
            min_shifts = min(weekend_shifts_assigned[name] for name in eligible)
            priority_pool = [name for name in eligible if weekend_shifts_assigned[name] == min_shifts]
            random.shuffle(priority_pool)
            secondary_pool = [name for name in eligible if weekend_shifts_assigned[name] > min_shifts]
            random.shuffle(secondary_pool)
            combined_pool = priority_pool + secondary_pool
            
            # 주말 근무 인원수가 부족해도 설정된 WEEKEND_STAFF_COUNT를 넘지 않음
            scheduled = combined_pool[:WEEKEND_STAFF_COUNT] 
            
            for name in scheduled:
                schedule_data[name].add(weekend_date.day)
                work_days_scheduled_this_week[name] += 1
                weekend_shifts_assigned[name] += 1

        # --- 평일 근무 배정 (균등 분배 로직으로 변경) ---
        
        # 1. 각 직원별로 남은 평일 근무 필요일수 계산
        remaining_weekday_shifts = {
            name: work_days_required[name] - work_days_scheduled_this_week[name]
            for name in all_employee_names
        }
        
        # 2. 요일별 근무 인원수 추적
        weekday_headcount = {day: 0 for day in weekdays_in_week}
        
        # 3. 모든 평일 근무가 배정될 때까지 반복
        total_shifts_to_assign = sum(remaining_weekday_shifts.values())
        
        for _ in range(total_shifts_to_assign):
            # 근무가 필요한 직원 목록
            employees_needing_shifts = [name for name, count in remaining_weekday_shifts.items() if count > 0]
            if not employees_needing_shifts: break

            # 가장 인원이 적은 평일(들)을 찾음
            min_headcount = min(weekday_headcount.values())
            least_staffed_days = [day for day, count in weekday_headcount.items() if count == min_headcount]
            
            # 그 중 하나를 무작위로 선택
            if not least_staffed_days: break
            target_day = random.choice(least_staffed_days)
            
            # 해당 날짜에 아직 배정되지 않았고, 근무가 필요한 직원들을 후보로 선정
            eligible_for_day = [
                name for name in employees_needing_shifts
                if target_day.day not in schedule_data[name]
            ]
            
            if not eligible_for_day:
                # 만약 해당 날에 채울 사람이 없으면 다른 비어있는 날을 시도
                random.shuffle(weekdays_in_week)
                found_slot = False
                for day in weekdays_in_week:
                    eligible_for_day = [name for name in employees_needing_shifts if day.day not in schedule_data[name]]
                    if eligible_for_day:
                        target_day = day
                        found_slot = True
                        break
                if not found_slot: continue

            # 최종 후보 중 한 명을 무작위로 선택하여 배정
            chosen_one = random.choice(eligible_for_day)
            
            schedule_data[chosen_one].add(target_day.day)
            remaining_weekday_shifts[chosen_one] -= 1
            weekday_headcount[target_day] += 1
            work_days_scheduled_this_week[chosen_one] += 1

    return schedule_data

def save_schedule_to_excel(employees_data, schedule_data, year, month):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년 {month}월 스케줄"
    output_filename = f"{year}년 {month}월 근무표.xlsx"
    _, num_days = calendar.monthrange(year, month)
    info_headers = ['contract_type', 'name', 'hire_date', 'contract_period', 'memo'] 
    display_info_headers = ['계약', '성명', '입사일', '계약만료일', '메모']
    summary_headers = ["평일 근무", "주말 근무"]
    last_col_num = len(info_headers) + num_days + len(summary_headers)
    title_text = f"{year}년 {month}월 근무표"
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=last_col_num)
    title_cell = ws.cell(row=2, column=1, value=title_text)
    title_cell.font = TITLE_FONT_EXCEL
    title_cell.alignment = CENTER_ALIGN
    for col, header_name in enumerate(display_info_headers, 1):
        cell = ws.cell(row=START_ROW, column=col, value=header_name)
        cell.fill = HEADER_FILL
        cell.font = BOLD_FONT_EXCEL
        cell.alignment = CENTER_ALIGN
    date_col_start = len(info_headers) + 1
    weekdays = ["월", "화", "수", "목", "금", "토", "일"]
    for day in range(1, num_days + 1):
        date = datetime.date(year, month, day)
        col = date_col_start + day - 1
        day_cell = ws.cell(row=START_ROW, column=col, value=f"{day}\n{weekdays[date.weekday()]}")
        day_cell.fill = HEADER_FILL
        day_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if date.weekday() == 5: day_cell.font = BLUE_FONT_EXCEL
        elif date.weekday() == 6: day_cell.font = RED_FONT_EXCEL
    for i, person_data in enumerate(employees_data):
        row_num = START_ROW + 1 + i
        for col, key in enumerate(info_headers, 1):
            ws.cell(row=row_num, column=col, value=person_data[key])
        work_days = schedule_data.get(person_data['name'], set())
        for day in work_days:
            col_num = date_col_start + day - 1
            ws.cell(row=row_num, column=col_num).fill = WORK_FILL
    summary_col_start = date_col_start + num_days
    for i, header in enumerate(summary_headers, 0):
        cell = ws.cell(row=START_ROW, column=summary_col_start + i, value=header)
        cell.fill = HEADER_FILL
        cell.font = BOLD_FONT_EXCEL
        cell.alignment = CENTER_ALIGN
    for i, person_data in enumerate(employees_data):
        row_num = START_ROW + 1 + i
        weekday_count, weekend_count = 0, 0
        work_days = schedule_data.get(person_data['name'], set())
        for day in work_days:
            if datetime.date(year, month, day).weekday() < 5: weekday_count += 1
            else: weekend_count += 1
        ws.cell(row=row_num, column=summary_col_start, value=weekday_count).alignment = CENTER_ALIGN
        ws.cell(row=row_num, column=summary_col_start + 1, value=weekend_count).alignment = CENTER_ALIGN
    last_row_num = START_ROW + len(employees_data)
    for r in ws.iter_rows(min_row=START_ROW, max_row=last_row_num, min_col=1, max_col=last_col_num):
        for cell in r:
            cell.border = TABLE_BORDER
    try:
        wb.save(output_filename)
        return f"✅ 성공! '{output_filename}' 파일이 생성되었습니다."
    except PermissionError:
        return f"❌ 실패: '{output_filename}' 파일이 다른 프로그램에서 열려 있습니다. 파일을 닫고 다시 시도해주세요."


# --- 3. Tkinter GUI 애플리케이션 (디자인 및 기능 개선) ---
class ScheduleApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("직원 근무 스케줄 관리")
        self.geometry("1100x750")

        self.employee_list = []
        self.selected_employee_id = None
        self.current_schedule_data = None
        self.contract_type_var = tk.StringVar(self) # 계약 형태 저장 변수
        
        self._create_widgets()
        self._apply_custom_styles()
        self._load_employees()

    def _apply_custom_styles(self):
        style = ttk.Style(self)
        style.configure("Accent.TButton", font=("맑은 고딕", 10, "bold"))
        try:
            self.app_icon = ImageTk.PhotoImage(file="app_icon.png")
            self.iconphoto(True, self.app_icon)
        except Exception as e:
            self._log_status(f"창 아이콘 로드 실패: {e}")

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_label = ttk.Label(main_frame, text="직원 근무 스케줄 관리 시스템", font=("맑은 고딕", 18, "bold"))
        title_label.pack(pady=(0, 20))

        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.BOTH, expand=True)

        list_frame = ttk.LabelFrame(top_frame, text="직원 목록", padding="10")
        list_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        self.employee_listbox = tk.Listbox(list_frame, height=25, width=30, font=("맑은 고딕", 10))
        self.employee_listbox.pack(side=tk.LEFT, fill=tk.Y)
        self.employee_listbox.bind('<<ListboxSelect>>', self._on_employee_select)
        list_scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.employee_listbox.yview)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.employee_listbox.config(yscrollcommand=list_scrollbar.set)

        form_frame = ttk.LabelFrame(top_frame, text="직원 정보", padding="15")
        form_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        form_frame.grid_columnconfigure(1, weight=1) # 컬럼 확장 설정

        # 3. 계약 버튼
        ttk.Label(form_frame, text="계약:").grid(row=0, column=0, sticky="nw", pady=5)
        contract_btn_frame = ttk.Frame(form_frame)
        contract_btn_frame.grid(row=0, column=1, sticky="ew", pady=5)
        ttk.Button(contract_btn_frame, text="주 2회", command=lambda: self._set_contract_type("주 2회")).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(contract_btn_frame, text="주 4회", command=lambda: self._set_contract_type("주 4회")).pack(side=tk.LEFT, padx=5)
        ttk.Button(contract_btn_frame, text="주 5회", command=lambda: self._set_contract_type("주 5회")).pack(side=tk.LEFT, padx=5)
        self.selected_contract_label = ttk.Label(contract_btn_frame, text="선택됨: 없음", font=("맑은 고딕", 10, "italic"), foreground="blue")
        self.selected_contract_label.pack(side=tk.LEFT, padx=10)
        
        ttk.Label(form_frame, text="성명:").grid(row=1, column=0, sticky="w", pady=5)
        self.name_entry = ttk.Entry(form_frame, width=35, font=("맑은 고딕", 10))
        self.name_entry.grid(row=1, column=1, sticky="ew", pady=5)

        # 4. DatePicker 적용
        ttk.Label(form_frame, text="입사일:").grid(row=2, column=0, sticky="w", pady=5)
        self.hire_date_entry = DateEntry(form_frame, width=32, background='darkblue', foreground='white', borderwidth=2,
                                         date_pattern='y-mm-dd', locale='ko_KR')
        self.hire_date_entry.grid(row=2, column=1, sticky="ew", pady=5)

        ttk.Label(form_frame, text="계약만료일:").grid(row=3, column=0, sticky="w", pady=5)
        self.contract_period_entry = DateEntry(form_frame, width=32, background='darkblue', foreground='white', borderwidth=2,
                                               date_pattern='y-mm-dd', locale='ko_KR')
        self.contract_period_entry.grid(row=3, column=1, sticky="ew", pady=5)

        ttk.Label(form_frame, text="메모:").grid(row=4, column=0, sticky="w", pady=5)
        self.memo_entry = ttk.Entry(form_frame, width=35, font=("맑은 고딕", 10))
        self.memo_entry.grid(row=4, column=1, sticky="ew", pady=5)

        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=5, column=0, columnspan=2, pady=15)
        # ... (버튼들)
        ttk.Button(button_frame, text="추가", command=self._add_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="수정", command=self._update_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="삭제", command=self._delete_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="초기화", command=self._clear_form).pack(side=tk.LEFT, padx=5)

        # 2. 상태 메시지 위치 변경
        status_frame = ttk.LabelFrame(form_frame, text="상태 메시지", padding="10")
        status_frame.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(10,0))
        self.status_text = scrolledtext.ScrolledText(status_frame, height=8, width=50, state=tk.DISABLED, font=("맑은 고딕", 9))
        self.status_text.pack(fill=tk.BOTH, expand=True)

        schedule_frame = ttk.LabelFrame(top_frame, text="스케줄 생성", padding="15")
        schedule_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        # ... (스케줄 생성 위젯들)
        ttk.Label(schedule_frame, text="연도:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.year_var = tk.StringVar(self)
        current_year = datetime.datetime.now().year
        self.year_combo = ttk.Combobox(schedule_frame, textvariable=self.year_var, values=[str(y) for y in range(current_year - 2, current_year + 5)], width=10, font=("맑은 고딕", 10))
        self.year_combo.set(str(current_year))
        self.year_combo.grid(row=0, column=1, sticky=tk.EW, pady=5)
        ttk.Label(schedule_frame, text="월:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.month_var = tk.StringVar(self)
        self.month_combo = ttk.Combobox(schedule_frame, textvariable=self.month_var, values=[str(m) for m in range(1, 13)], width=10, font=("맑은 고딕", 10))
        self.month_combo.set(str(datetime.datetime.now().month))
        self.month_combo.grid(row=1, column=1, sticky=tk.EW, pady=5)
        ttk.Button(schedule_frame, text="스케줄 미리보기 및 생성", command=self._show_preview_window, style="Accent.TButton").grid(row=2, column=0, columnspan=2, pady=20)


    def _show_preview_window(self):
        try:
            year = int(self.year_var.get())
            month = int(self.month_var.get())
        except ValueError:
            messagebox.showwarning("입력 오류", "연도와 월을 올바르게 선택해주세요.")
            return

        self.current_employees_data = get_all_employees_db()
        if not self.current_employees_data: return messagebox.showwarning("경고", "데이터베이스에 직원이 없습니다.")

        preview_window = tk.Toplevel(self)
        preview_window.title(f"{year}년 {month}월 스케줄 미리보기")
        
        # 1. 미리보기 헤더에 요일 추가
        _, num_days = calendar.monthrange(year, month)
        weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        cols = ['성명', '계약'] + [f"{d} ({weekdays[datetime.date(year, month, d).weekday()]})" for d in range(1, num_days + 1)]
        tree = ttk.Treeview(preview_window, columns=cols, show='headings')
        
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=45, anchor=tk.CENTER)
        tree.column('성명', width=80)
        tree.column('계약', width=60)
        tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        def populate_tree():
            for i in tree.get_children(): tree.delete(i)
            self.current_schedule_data = create_schedule_data(self.current_employees_data, year, month)
            for emp in self.current_employees_data:
                row_data = [emp['name'], emp['contract_type']]
                work_days = self.current_schedule_data.get(emp['name'], set())
                for day in range(1, num_days + 1):
                    row_data.append("O" if day in work_days else "")
                tree.insert("", "end", values=row_data)
            self._log_status("스케줄 미리보기를 다시 생성했습니다.")
        
        def save_and_close():
            result_message = save_schedule_to_excel(self.current_employees_data, self.current_schedule_data, year, month)
            self._log_status(result_message)
            messagebox.showinfo("스케줄 생성 결과", result_message)
            preview_window.destroy()

        preview_button_frame = ttk.Frame(preview_window, padding="5")
        preview_button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)
        # ... (미리보기 버튼들)
        ttk.Button(preview_button_frame, text="다시 생성", command=populate_tree).pack(side=tk.LEFT, expand=True, padx=5)
        try:
            self.save_icon = ImageTk.PhotoImage(Image.open("save_icon.png").resize((16, 16)))
            ttk.Button(preview_button_frame, text="이대로 엑셀 저장", image=self.save_icon, compound="left", command=save_and_close, style="Accent.TButton").pack(side=tk.LEFT, expand=True, padx=5)
        except Exception:
            ttk.Button(preview_button_frame, text="이대로 엑셀 저장", command=save_and_close, style="Accent.TButton").pack(side=tk.LEFT, expand=True, padx=5)
        ttk.Button(preview_button_frame, text="취소", command=preview_window.destroy).pack(side=tk.LEFT, expand=True, padx=5)
        
        populate_tree()

    def _set_contract_type(self, contract_type):
        """계약 버튼 클릭 시 호출되는 함수"""
        self.contract_type_var.set(contract_type)
        self.selected_contract_label.config(text=f"선택됨: {contract_type}")

    def _on_employee_select(self, event):
        selected_indices = self.employee_listbox.curselection()
        if selected_indices:
            index = selected_indices[0]
            selected_employee = self.employee_list[index]
            self.selected_employee_id = selected_employee['id']
            
            # 계약, 성명, 메모 설정
            self._set_contract_type(selected_employee['contract_type'])
            self.name_entry.delete(0, tk.END); self.name_entry.insert(0, selected_employee['name'])
            self.memo_entry.delete(0, tk.END); self.memo_entry.insert(0, selected_employee['memo'])
            
            # DatePicker 값 설정
            try:
                if selected_employee['hire_date']:
                    self.hire_date_entry.set_date(datetime.datetime.strptime(selected_employee['hire_date'], '%Y-%m-%d'))
                if selected_employee['contract_period']:
                    self.contract_period_entry.set_date(datetime.datetime.strptime(selected_employee['contract_period'], '%Y-%m-%d'))
            except (ValueError, TypeError):
                self._log_status("DB의 날짜 형식이 잘못되었습니다. (YYYY-MM-DD)")
        else: self._clear_form()
    
    def _clear_form(self):
        self.selected_employee_id = None
        self._set_contract_type("") # 계약 정보 초기화
        self.selected_contract_label.config(text="선택됨: 없음")
        self.name_entry.delete(0, tk.END)
        self.memo_entry.delete(0, tk.END)
        self.hire_date_entry.set_date(datetime.date.today()) # 오늘 날짜로 초기화
        self.contract_period_entry.set_date(datetime.date.today())
        self.employee_listbox.selection_clear(0, tk.END)
        self._log_status("직원 정보 입력 폼이 초기화되었습니다.")

    def _get_form_data(self):
        contract_type = self.contract_type_var.get()
        name = self.name_entry.get().strip()
        hire_date = self.hire_date_entry.get()
        contract_period = self.contract_period_entry.get()
        memo = self.memo_entry.get().strip()
        if not name or not contract_type:
            messagebox.showwarning("입력 오류", "계약 형태와 성명은 필수 입력 항목입니다.")
            return None
        return (contract_type, name, hire_date, contract_period, memo)
    
    # --- 나머지 GUI 메소드들 (이전과 동일) ---
    def _log_status(self, message):
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.status_text.see(tk.END)
        self.status_text.config(state=tk.DISABLED)
    def _load_employees(self):
        self.employee_listbox.delete(0, tk.END)
        self.employee_list = get_all_employees_db()
        for i, emp in enumerate(self.employee_list):
            self.employee_listbox.insert(tk.END, f"{emp['name']} ({emp['contract_type']})")
        self._log_status("직원 목록을 불러왔습니다.")
    def _add_employee(self):
        data = self._get_form_data()
        if data and add_employee_db(*data):
            self._log_status(f"{data[1]}님 정보가 성공적으로 추가되었습니다.")
            self._clear_form()
            self._load_employees()
    def _update_employee(self):
        if not self.selected_employee_id: messagebox.showwarning("선택 오류", "수정할 직원을 목록에서 선택해주세요."); return
        data = self._get_form_data()
        if data and update_employee_db(self.selected_employee_id, *data):
            self._log_status(f"{data[1]}님 정보가 성공적으로 수정되었습니다.")
            self._clear_form()
            self._load_employees()
    def _delete_employee(self):
        if not self.selected_employee_id: messagebox.showwarning("선택 오류", "삭제할 직원을 목록에서 선택해주세요."); return
        if messagebox.askyesno("삭제 확인", f"'{self.name_entry.get()}' 님을 정말 삭제하시겠습니까?"):
            if delete_employee_db(self.selected_employee_id):
                self._log_status(f"{self.name_entry.get()}님 정보가 성공적으로 삭제되었습니다.")
                self._clear_form()
                self._load_employees()

# --- 메인 실행 ---
if __name__ == "__main__":
    init_db()
    app = ScheduleApp()
    app.mainloop()